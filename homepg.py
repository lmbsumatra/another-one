import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from balanceinq import balanceInq


def homePage():
    option = '''    [1] Balance Inquire
    [2] Deposit
    [3] Withdraw'''
    print(option)

    # check_acctno = 987456321
    # bal_inq = 100
    # count = 0
    # check_bal = 0

    # choice = int(input("Enter your choice: "))
    # if choice == 1:
    #     balanceInq(check_acctno, check_bal)
# trial uli for acct no and balance
def getData():
    check_acctno = 987456321
    bal_inq = 100
    count = 0
    check_bal = 0

    # creating file for trial hoho
    # wb = Workbook()
    # wb.save(filename = 'nag-iisa-lang-to.xlsx')

    

    xl = openpyxl.load_workbook("nag-iisa-lang-to.xlsx")
  
    data = xl.active

    
    acctno_column = data['A']           # get data from A column
    
    # parang may ibang way pa na mas maganda to find data? ayaw lang i-type ng daliri ko 
    # for x in range(len(acctno_column)):
    #     if acct_no == acctno_column[x].value:
    #         print("hello")
    #     count = count + 1 

    # for row in data.iter_rows(min_row = count, max_col= None, max_row = count, values_only = True):
    #     for infos in row:
    #         if infos == bal_inq:
    #             print("korik dzai")
    # print(count)
    # for row in data.iter_rows(min_row = 1, min_col = 1, max_row = None, max_col = None):
    #     for cell in row:
    #         if cell.value == 987456321:
    #             print(data.cell(row=cell.row, column=cell.column))
    #         count = count + 1 
    #         print(count)
    # for row in data.iter_rows(min_row = count, max_col= None, max_row = count, values_only = True):
    #     for infos in row:
    #         if infos == bal_inq:
    #             kkk = infos
    #             print(infos)

    

homePage()